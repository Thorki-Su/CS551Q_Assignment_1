import os
from pathlib import Path
from django.db import models
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from visual_emission.models import Country, Data # load the models

class Command(BaseCommand):
    help = 'Load data from csv'

    def handle(self, *args, **options):
        Country.objects.all().delete()
        Data.objects.all().delete()
        print('table dropped')

        base_dir = Path(__file__).resolve().parent.parent.parent.parent
        book_path = os.path.join(base_dir, 'visual_emission/country_data/data_upload.xlsx')
        book = load_workbook(book_path)
        sheet = book['Data']
        print(sheet.title)
        max_row_num = sheet.max_row
        max_col_num = sheet.max_column
        print(f'Rows: {max_row_num}, Columns: {max_col_num}')

        start_year = 1990 # years are from 1990 to 2020
        data_start_col = 6 # data start at column F, which is the 6th

        for i in range(2, max_row_num + 1):
            row_data = [sheet.cell(row=i, column=j).value for j in range(1, max_col_num + 1)] # get the data of one row
            c_name = row_data[0]
            c_code = row_data[1]
            is_c = row_data[2]
            reg = row_data[3] if row_data[3] else ''
            income = row_data[4] if row_data[4] else ''

            country = Country.objects.create(
                country_name = c_name,
                country_code = c_code,
                is_country = is_c,
                region = reg,
                income_group = income,
            )

            data_objects = []
            for j in range(data_start_col-1, max_col_num): # the index starts with 0, so the data_start_col should -1
                year = start_year + (j-data_start_col+1)
                emission = row_data[j]
                if emission is not None:
                    data_objects.append(Data(country = country, year = year, emission = emission))
                
            Data.objects.bulk_create(data_objects)
            print(f'{c_name} saved')
        
        print('all data saved')